import pandas as pd
import os
import glob
import csv
from xlsxwriter.workbook import Workbook

cwd = os.getcwd()
users_dir = os.getcwd() + '/data/all'
data_file = 'all_data.xlsx'
xlsx_suffix = '.xlsx'
csv_suffix = '.csv'
output_dir = './output'
chunksize = 50000

columns = ['Albümin/Kreatinin (Spot idrar)', 'Demir (Serum/Plazma)',
		   'Ferritin (Serum/Plazma)',
		   'Glike hemoglobin (Hb A1c) (HPLC)', 'HbA1C',
		   'Hemoglobin (HGB) (Hemogram(Tam Kan))',
		   'Kreatinin (Kreatinin (Serum/Plazma))',
		   'UIBC']



def clean_invalid_chars(s):
	"""

	:param s:
	char to clean
	:return:
	Returns converted float value if it is possible to convert
	"""
	try:
		s = s.replace(',', '.')
		i = float(s)
		return i
	except ValueError:
		return 0
	except Exception:
		return 0


def get_file_names(directory):
	"""
	:param directory:
	dir to look at for files
	:return:
	Returns list of file names
	"""
	files_to_read = []
	for filename in os.listdir(directory):
		if filename.endswith(xlsx_suffix):
			files_to_read.append(os.path.join(directory, filename))
		else:
			continue
	return files_to_read


def clean_dataframe(df):
	"""

	:param df:
	Dataframe to clean
	:return:
	Dataframe which is cleaned for the purpose
	"""
	df = df[['Test Adı', 'Sonuç']]
	df = df.fillna(0)
	df['Sonuç'] = df['Sonuç'].apply(lambda x: clean_invalid_chars(x))
	df['Sonuç'] = df['Sonuç'].astype(float)
	return df


def get_average(df):
	"""

	:param df:
	:return:
	Returns mean of given dataframe grouped by column name
	"""
	df_mean = df.groupby('Test Adı', as_index=True).mean()
	d = df_mean.T
	user_value = d[d.columns.intersection(columns)]
	f = user_value.rename(columns={'Test Adı': 'Isim Soyisim'}, index={'Sonuç': filename.split(xlsx_suffix)[0]})
	return f


def format_values(df, columns=columns):
	"""
	:param df:
	:param columns:
	"""
	df = df[['Test Adı', 'Sonuç']]
	df.set_index('Test Adı', inplace=True)
	nd = df.T
	# user_value = nd[nd.columns.intersection(columns)]
	return nd


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
					   truncate_sheet=False,
					   **to_excel_kwargs):
	"""
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
	from openpyxl import load_workbook

	# ignore [engine] parameter if it was passed
	if 'engine' in to_excel_kwargs:
		to_excel_kwargs.pop('engine')

	writer = pd.ExcelWriter(filename, engine='openpyxl')

	try:
		# try to open an existing workbook
		writer.book = load_workbook(filename)

		# get the last row in the existing Excel sheet
		# if it was not specified explicitly
		if startrow is None and sheet_name in writer.book.sheetnames:
			startrow = writer.book[sheet_name].max_row

		# truncate sheet
		if truncate_sheet and sheet_name in writer.book.sheetnames:
			# index of [sheet_name] sheet
			idx = writer.book.sheetnames.index(sheet_name)
			# remove [sheet_name]
			writer.book.remove(writer.book.worksheets[idx])
			# create an empty sheet [sheet_name] using old index
			writer.book.create_sheet(sheet_name, idx)

		# copy existing sheets
		writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
	except FileNotFoundError:
		# file does not exist yet, we will create it
		pass

	if startrow is None:
		startrow = 0

	# write out the new sheet
	df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

	# save the workbook
	writer.save()


def create_sep_files(columns, files):
	for c in columns:
		for f in files:
			chunks = f.split('/')
			filename = chunks[7]
			try:
				user_data = pd.read_excel(f)
			except ValueError:
				pass
			except Exception:
				pass
			df = clean_dataframe(user_data)
			df = format_values(user_data)
			df = df.rename(columns={'Test Adı': 'Isim Soyisim'}, index={'Sonuç': filename.split(xlsx_suffix)[0]})
			try:
				data = df[c]
				write_to_csv(data, c)
			except KeyError:
				data[c] = 'Nan'
				print('KeyError: {} does not exist in the table.'.format(c))
				continue


def write_to_csv(data, c):
	"""

	:param data:
	:param c:
	"""
	if not os.path.isfile('./output/{}.csv'.format(c.replace('/','_'))):
		data.to_csv('./output/{}.csv'.format(c.replace('/','_')), header=c, encoding='utf-8-sig')
	else:  # else it exists so append without writing the header
		data.to_csv('./output/{}.csv'.format(c.replace('/','_')), mode='a', header=False, encoding='utf-8-sig')


def csv_to_xlsx(csv_dir):
	"""

	:param csv_dir:
	"""
	for csvfile in glob.glob(os.path.join(csv_dir, '*.csv')):
		workbook = Workbook(csvfile[:-4] + '.xlsx')
		worksheet = workbook.add_worksheet()
		with open(csvfile, 'rt', encoding='utf8') as f:
			reader = csv.reader(f)
			for r, row in enumerate(reader):
				for c, col in enumerate(row):
					worksheet.write(r, c, col)
		workbook.close()


def remove_csv_files(dir):
	"""
	:param dir:
	"""
	for csvfile in glob.glob(os.path.join(dir, '*.csv')):
		os.remove(csvfile)


# def append_test_name():

if __name__ == '__main__':
	files = get_file_names(output_dir)
	# create_sep_files(columns,files)
	# csv_to_xlsx('./output')
	# remove_csv_files('./output')
	# files = get_file_names(output_dir)
	data_frames = []
	for f in files:
		chunks = f.split('/')
		filename = chunks[2]
		try:
			user_data = pd.read_excel(f)
			u = user_data.reset_index()
			u.set_index('level_0', inplace=True)

			l = len(u.columns)

			liste = [filename.split('.xlsx')[0].replace('_','/') for x in range(l)]
			u.columns = liste
		except ValueError:
			pass
		except Exception:
			pass
		data_frames.append(u)

	# frame = pd.DataFrame()
	# result = frame.append(data_frames)
	# main_frame= pd.concat([df.stack() for df in data_frames], axis=0).unstack()
	main_frame = pd.concat( data_frames,axis=1 ,sort=False)
	main_frame.to_excel('./output/all.xlsx')
	# create_sep_files(columns, files)
	# csv_to_xlsx(output_dir)
	# remove_csv_files(output_dir)